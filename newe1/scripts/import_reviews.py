#!/usr/bin/env python3
"""
Script to import reviews data from Excel file to PostgreSQL database.
"""

import os
import sys
import psycopg2
from psycopg2.extras import execute_values
import openpyxl
from datetime import datetime

# Database connection parameters
DB_CONFIG = {
    'host': os.environ.get('POSTGRES_HOST', '192.168.40.41'),
    'port': int(os.environ.get('POSTGRES_PORT', '5432')),
    'database': os.environ.get('POSTGRES_DB', 'newe1'),
    'user': os.environ.get('POSTGRES_USER', 'newe1'),
    'password': os.environ.get('POSTGRES_PASSWORD', 'newe1pass'),
}

# SQL to create the reviews table
CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS reviews (
    id SERIAL PRIMARY KEY,
    original_id INTEGER UNIQUE,
    name VARCHAR(255) NOT NULL,
    review_text TEXT NOT NULL,
    company_response TEXT,
    rating INTEGER CHECK (rating >= 1 AND rating <= 5),
    is_active BOOLEAN DEFAULT true,
    show_on_main BOOLEAN DEFAULT false,
    photos TEXT[],
    order_number VARCHAR(100),
    phone VARCHAR(50),
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_reviews_active ON reviews(is_active);
CREATE INDEX IF NOT EXISTS idx_reviews_created_at ON reviews(created_at DESC);
CREATE INDEX IF NOT EXISTS idx_reviews_show_on_main ON reviews(show_on_main);
"""

def parse_date(date_str):
    """Parse date string like '15.01.2024 13:42:15' to datetime."""
    if not date_str:
        return None
    try:
        return datetime.strptime(str(date_str).strip(), '%d.%m.%Y %H:%M:%S')
    except (ValueError, AttributeError):
        try:
            return datetime.strptime(str(date_str).strip(), '%d.%m.%Y')
        except (ValueError, AttributeError):
            return None

def parse_rating(rating_str):
    """Parse rating from string or number."""
    if not rating_str:
        return None
    try:
        rating = int(str(rating_str).strip())
        return rating if 1 <= rating <= 5 else None
    except (ValueError, TypeError):
        return None

def parse_bool(value):
    """Parse boolean from various formats."""
    if not value:
        return False
    return str(value).strip().lower() in ('да', 'yes', 'true', '1')

def connect_db():
    """Connect to PostgreSQL database."""
    return psycopg2.connect(**DB_CONFIG)

def create_table(conn):
    """Create the reviews table if it doesn't exist."""
    with conn.cursor() as cur:
        cur.execute(CREATE_TABLE_SQL)
    conn.commit()
    print("Table 'reviews' created/verified successfully")

def import_data(conn, excel_path):
    """Import data from Excel file to database."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    print(f"Total rows: {ws.max_row}")

    # Map column indices (0-indexed)
    # Headers: ['Название', 'Активность', 'Дата изменения', 'Дата создания',
    #           'Описание для анонса', 'Ответ на отзыв', 'ID',
    #           'Ответивший сотрудник', 'Показывать на главной', 'Рейтинг']
    col_map = {
        'name': 0,           # Название (ФИО)
        'is_active': 1,      # Активность
        'updated_at': 2,     # Дата изменения
        'created_at': 3,     # Дата создания
        'review_text': 4,    # Описание для анонса (текст отзыва)
        'company_response': 5,  # Ответ на отзыв
        'original_id': 6,    # ID
        'show_on_main': 8,   # Показывать на главной
        'rating': 9,         # Рейтинг
    }

    records = []
    for row_num in range(2, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

        name = row[col_map['name']]
        review_text = row[col_map['review_text']]

        # Skip rows without name or review text
        if not name or not review_text:
            continue

        # Get original ID
        original_id = row[col_map['original_id']]
        try:
            original_id = int(original_id) if original_id else None
        except (ValueError, TypeError):
            original_id = None

        record = {
            'name': str(name).strip(),
            'review_text': str(review_text).strip(),
            'company_response': row[col_map['company_response']] or None,
            'rating': parse_rating(row[col_map['rating']]),
            'is_active': parse_bool(row[col_map['is_active']]),
            'show_on_main': parse_bool(row[col_map['show_on_main']]),
            'created_at': parse_date(row[col_map['created_at']]),
            'updated_at': parse_date(row[col_map['updated_at']]),
            'original_id': original_id,
        }

        records.append(record)

    # Insert records using execute_values for efficiency
    insert_sql = """
        INSERT INTO reviews (name, review_text, company_response, rating,
                            is_active, show_on_main, created_at, updated_at, original_id)
        VALUES %s
        ON CONFLICT (original_id) DO UPDATE SET
            name = EXCLUDED.name,
            review_text = EXCLUDED.review_text,
            company_response = EXCLUDED.company_response,
            rating = EXCLUDED.rating,
            is_active = EXCLUDED.is_active,
            show_on_main = EXCLUDED.show_on_main,
            updated_at = EXCLUDED.updated_at
    """

    values = [
        (r['name'], r['review_text'], r['company_response'], r['rating'],
         r['is_active'], r['show_on_main'], r['created_at'], r['updated_at'], r['original_id'])
        for r in records
    ]

    with conn.cursor() as cur:
        execute_values(cur, insert_sql, values)
    conn.commit()

    print(f"Imported {len(records)} reviews successfully")

def main():
    # Get Excel file path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, '..', 'uploads', 'otzivi.xlsx')

    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)

    print(f"Connecting to database at {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}")

    try:
        conn = connect_db()
        print("Connected to database successfully")

        create_table(conn)
        import_data(conn, excel_path)

        # Verify import
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM reviews")
            count = cur.fetchone()[0]
            print(f"Total reviews in database: {count}")

            cur.execute("SELECT COUNT(*) FROM reviews WHERE is_active = true")
            active_count = cur.fetchone()[0]
            print(f"Active reviews: {active_count}")

            cur.execute("SELECT AVG(rating)::numeric(2,1) FROM reviews WHERE rating IS NOT NULL")
            avg_rating = cur.fetchone()[0]
            print(f"Average rating: {avg_rating}")

        conn.close()

    except psycopg2.Error as e:
        print(f"Database error: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
