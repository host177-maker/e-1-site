#!/usr/bin/env python3
"""
Script to import salon data from Excel file to PostgreSQL database.
"""

import os
import sys
import re
import psycopg2
from psycopg2.extras import execute_values
import openpyxl

# Database connection parameters (все значения из env переменных)
if not os.environ.get('POSTGRES_HOST'):
    print("❌ POSTGRES_HOST не установлен")
    sys.exit(1)

DB_CONFIG = {
    'host': os.environ.get('POSTGRES_HOST'),
    'port': int(os.environ.get('POSTGRES_PORT', '5432')),
    'database': os.environ.get('POSTGRES_DB'),
    'user': os.environ.get('POSTGRES_USER'),
    'password': os.environ.get('POSTGRES_PASSWORD'),
}

# SQL to create the salons table
CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS salons (
    id SERIAL PRIMARY KEY,
    original_id INTEGER,
    name VARCHAR(255) NOT NULL,
    city VARCHAR(100) NOT NULL,
    region VARCHAR(100),
    address TEXT,
    email VARCHAR(255),
    phone VARCHAR(50),
    working_hours VARCHAR(255),
    latitude DECIMAL(10, 6),
    longitude DECIMAL(10, 6),
    external_code VARCHAR(50),
    slug VARCHAR(255),
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(original_id)
);

CREATE INDEX IF NOT EXISTS idx_salons_city ON salons(city);
CREATE INDEX IF NOT EXISTS idx_salons_region ON salons(region);
"""

def parse_coordinates(coord_str):
    """Parse coordinates string like '45.101305,38.983835' into lat, lng."""
    if not coord_str:
        return None, None
    try:
        parts = coord_str.strip().split(',')
        if len(parts) == 2:
            lat = float(parts[0].strip())
            lng = float(parts[1].strip())
            return lat, lng
    except (ValueError, AttributeError):
        pass
    return None, None

def parse_region(region_str):
    """Parse region string like 'Краснодар [3481]' to extract city name."""
    if not region_str:
        return None
    # Remove bracketed ID
    return re.sub(r'\s*\[\d+\]$', '', region_str).strip()

def clean_working_hours(hours_str):
    """Clean working hours string, removing HTML: prefix."""
    if not hours_str:
        return None
    # Remove HTML: prefix
    hours = re.sub(r'^HTML:', '', str(hours_str)).strip()
    return hours if hours else None

def connect_db():
    """Connect to PostgreSQL database."""
    return psycopg2.connect(**DB_CONFIG)

def create_table(conn):
    """Create the salons table if it doesn't exist."""
    with conn.cursor() as cur:
        cur.execute(CREATE_TABLE_SQL)
    conn.commit()
    print("Table 'salons' created/verified successfully")

def import_data(conn, excel_path):
    """Import data from Excel file to database."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    print(f"Total rows: {ws.max_row}")

    # Map column indices (1-indexed)
    col_map = {
        'name': 1,           # Название
        'city': 2,           # Разделы (это город)
        'slug': 5,           # Символьный код
        'external_code': 6,  # Внешний код
        'original_id': 7,    # ID
        'email': 8,          # E-mail
        'address': 9,        # Адрес
        'coordinates': 10,   # Координаты на карте
        'region': 11,        # Регион
        'working_hours': 12, # Режим работы
        'phone': 13,         # Телефон
    }

    records = []
    for row_num in range(2, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

        # Parse coordinates
        coord_str = row[col_map['coordinates'] - 1] if len(row) >= col_map['coordinates'] else None
        lat, lng = parse_coordinates(coord_str)

        # Parse region (extract city name without ID)
        region_str = row[col_map['region'] - 1] if len(row) >= col_map['region'] else None
        region = parse_region(region_str)

        # Clean working hours
        working_hours_str = row[col_map['working_hours'] - 1] if len(row) >= col_map['working_hours'] else None
        working_hours = clean_working_hours(working_hours_str)

        # Get original ID
        original_id = row[col_map['original_id'] - 1] if len(row) >= col_map['original_id'] else None
        try:
            original_id = int(original_id) if original_id else None
        except (ValueError, TypeError):
            original_id = None

        record = {
            'name': row[col_map['name'] - 1],
            'city': row[col_map['city'] - 1],
            'region': region or row[col_map['city'] - 1],  # Use city as region if region is empty
            'address': row[col_map['address'] - 1],
            'email': row[col_map['email'] - 1],
            'phone': row[col_map['phone'] - 1],
            'working_hours': working_hours,
            'latitude': lat,
            'longitude': lng,
            'external_code': row[col_map['external_code'] - 1],
            'slug': row[col_map['slug'] - 1],
            'original_id': original_id,
        }

        records.append(record)

    # Insert records using execute_values for efficiency
    insert_sql = """
        INSERT INTO salons (name, city, region, address, email, phone, working_hours,
                           latitude, longitude, external_code, slug, original_id)
        VALUES %s
        ON CONFLICT (original_id) DO UPDATE SET
            name = EXCLUDED.name,
            city = EXCLUDED.city,
            region = EXCLUDED.region,
            address = EXCLUDED.address,
            email = EXCLUDED.email,
            phone = EXCLUDED.phone,
            working_hours = EXCLUDED.working_hours,
            latitude = EXCLUDED.latitude,
            longitude = EXCLUDED.longitude,
            external_code = EXCLUDED.external_code,
            slug = EXCLUDED.slug
    """

    values = [
        (r['name'], r['city'], r['region'], r['address'], r['email'], r['phone'],
         r['working_hours'], r['latitude'], r['longitude'], r['external_code'],
         r['slug'], r['original_id'])
        for r in records
    ]

    with conn.cursor() as cur:
        execute_values(cur, insert_sql, values)
    conn.commit()

    print(f"Imported {len(records)} salons successfully")

def main():
    # Get Excel file path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, '..', 'uploads', 'salone1.xlsx')

    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)

    print(f"Connecting to database at {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}")

    try:
        conn = connect_db()
        print("Connected to database successfully")

        create_table(conn)
        import_data(conn, excel_path)

        # Verify import
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM salons")
            count = cur.fetchone()[0]
            print(f"Total salons in database: {count}")

            cur.execute("SELECT DISTINCT city FROM salons ORDER BY city")
            cities = [row[0] for row in cur.fetchall()]
            print(f"Cities: {', '.join(cities[:10])}{'...' if len(cities) > 10 else ''}")

        conn.close()

    except psycopg2.Error as e:
        print(f"Database error: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
